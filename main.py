import os
import argparse

from omegaconf import DictConfig
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.comments import Comment

from func_utils import *

args_parser = argparse.ArgumentParser(description="使用演變表一鍵轉換出字表    By EcisralHetha")
args_parser.add_argument('地名', type=str, help='地名，需要包括小地名在內的全名')
args_parser.add_argument('-i', '--輸入路徑', type=str, default='./泛粵字表 21.xlsx', help='泛粵表本體的路徑，默認 = "./泛粵字表 21.xlsx"')
args_parser.add_argument('-o', '--輸出路徑', type=str, default='./轉換字表.xlsx', help='輸出字表的路徑，默認 = "./轉換字表.xlsx"')
args_parser.add_argument('-k', '--地位路徑', type=str, default='./廣韻4k.xlsx', help='音韻地位字表的路徑，默認 = "./廣韻4k.xlsx"')
args_parser.add_argument('-A', help='全濁上【不變】去', action='store_false')
args_parser.add_argument('-B', help='次濁上【 變 】去', action='store_true')
args_parser.add_argument('-v', '--version', action='version', version='v0.5/211017')
args = args_parser.parse_args()

target_loc_name = args.地名
config_全濁上變去 = args.A
config_次濁上變去 = args.B
input_path_sheet = args.輸入路徑
output_sheet_name = args.輸出路徑
input_path_kwangyon = args.地位路徑

if not os.path.isfile(input_path_sheet): raise IOError("輸入字表不存在，請覈對字表路徑")
if not os.path.isfile(input_path_kwangyon): raise IOError("地位字表不存在，請覈對字表路徑")
if os.path.isfile(output_sheet_name): print("-- 記得關掉輸出表格 --")

#-----------------------------------------#

print("[1] 正在讀取主表")
pan_cantonese_sheet = load_workbook(filename=input_path_sheet, data_only=True)

#-----------------------------------------#

print("[2] 正在讀取地位")
KwangyonSheet = load_workbook(filename=input_path_kwangyon, data_only=True)
kwangyon_sheet = KwangyonSheet["Sheet1"]

chara_to_position = {}
for n,i in enumerate(kwangyon_sheet):
    if n==0: continue
    # ↓ 消去重紐標記，sheet_row[0] ==> 字頭
    sheet_row = [(j.value.replace("A", "").replace("B", "") if j.column==3 else j.value) if j.value else "" for j in i]
    if sheet_row[0] not in chara_to_position:
        chara_to_position[sheet_row[0]] = [sheet_row[1:]]
    else:
        chara_to_position[sheet_row[0]].append(sheet_row[1:])
# chara_to_position["重"] ==> [['澄','鍾','開','三','通','平'], ['澄','鍾','開','三','通','上'], ['澄','鍾','開','三','通','去']]

#-----------------------------------------#

print("[3] 正在處理")
final_sheet = pan_cantonese_sheet["演變表-韻母"]
target_loc_index = {(target_loc_name if i.value==target_loc_name else ""):i.column for i in final_sheet[1]}
target_loc_index = target_loc_index[target_loc_name]-1
sheet_header = ["攝","呼","等","韻","聲母","例字","另讀","廢音","例外"] # 代碼用不到，衹是提示表頭罷了
sheet_rows_header = []
sheet_last_row = ["_"] * 5
for i in range(3,final_sheet.max_row+1):
    sheet_row = []
    for j in range(len(sheet_header)):
        value = final_sheet[i][j].value.strip() if final_sheet[i][j].value else ""
        if j<5: # 中古音分化條件有 5 列，見 sheet_header
            if value!="":
                if j==4: value = expend_intial_expr(value)
                sheet_last_row[j] = value
                sheet_last_row[j+1:5] = ["_"] * len(sheet_last_row[j+1:5])
            elif j==4: sheet_last_row[j] = "_"
        else:
            sheet_row.append(value)
    sheet_row_content = final_sheet[i][target_loc_index].value.strip() if final_sheet[i][target_loc_index].value else ""
    sheet_rows_header.append(sheet_last_row + sheet_row + [sheet_row_content])
    
final_config = DictConfig(content={})
for n, i in enumerate(sheet_rows_header):
    v0, v1, v2, v3, v4 = i[:5]
    if v0 not in final_config: final_config[v0] = {}
    if v1 not in final_config[v0]: final_config[v0][v1] = {}
    if v2 not in final_config[v0][v1]: final_config[v0][v1][v2] = {}
    if v3 not in final_config[v0][v1][v2]: final_config[v0][v1][v2][v3] = {}
    final_config[v0][v1][v2][v3][v4] = [i[-1], i[6], i[7], i[8]]
# final_config.宕.三._.見系 ==> [oeng, 另讀, 廢音, 例外]

#-----------------------------------------#

initial_sheet = pan_cantonese_sheet["演變表-聲母"]
target_loc_index = {(target_loc_name if i.value==target_loc_name else ""):i.column for i in initial_sheet[1]}
target_loc_index = target_loc_index[target_loc_name]-1
sheet_header = ["聲","呼","等","今調","攝","韻","例字","另讀","清濁","廢音"] # 代碼用不到，衹是提示表頭罷了
sheet_rows_header = []
sheet_last_row = ["_"] * 6
for i in range(2,initial_sheet.max_row+1):
    sheet_row = []
    for j in range(1,len(sheet_header)+1): # 從 1 開始，因爲表格首列冇用，跳過
        value = initial_sheet[i][j].value.strip() if initial_sheet[i][j].value else ""
        if j<=6: # 中古音分化條件有 6 列，見 sheet_header
            if value!="":
                if j==1: value = expend_intial_expr(value)
                sheet_last_row[j-1] = value
                sheet_last_row[j:] = ["_"] * len(sheet_last_row[j:])
            elif j==5: sheet_last_row[j] = "_"
        else:
            sheet_row.append(value)
    sheet_row_content = initial_sheet[i][target_loc_index].value.strip().replace("0", "") if initial_sheet[i][target_loc_index].value else ""
    sheet_rows_header.append(sheet_last_row + sheet_row + [sheet_row_content])
# (sheet_last_row, sheet_row, sheet_row_content) ==> (['幫', '_', '_', '_', '_', '_'], ['般邦扁杯', '', '', '費俸'], 'b')

initial_config = DictConfig(content={})
for n, i in enumerate(sheet_rows_header):
    v0, v1, v2, v3, v4, v5 = i[:6]
    if v0 not in initial_config: initial_config[v0] = {}
    if v1 not in initial_config[v0]: initial_config[v0][v1] = {}
    if v2 not in initial_config[v0][v1]: initial_config[v0][v1][v2] = {}
    if v3 not in initial_config[v0][v1][v2]: initial_config[v0][v1][v2][v3] = {}
    if v4 not in initial_config[v0][v1][v2][v3]: initial_config[v0][v1][v2][v3][v4] = {}
    initial_config[v0][v1][v2][v3][v4][v5] = [i[-1], i[7], i[8], i[9]]
# initial_config.見.合.一.蟹.灰 ==> [gw, 另讀, 清濁不合, 廢音]

#-----------------------------------------#

tone_sheet = pan_cantonese_sheet["聲調總表"]
target_loc_index = {(target_loc_name if (i[0].value if i[0].value else "")+(i[1].value if i[1].value else "")==target_loc_name else ""):i[0].row for i in zip(tone_sheet["B"], tone_sheet["C"])}
target_loc_index = target_loc_index[target_loc_name]
# 下面的 3:44 是聲調表的 D 列到 AR 列，即十個傳統聲調 + 非傳統聲調部分
tone_class_contour = \
    [i.value if i.value else "" for n,i in enumerate(tone_sheet[target_loc_index][3:44:3])][:10] + \
    [i.value if i.value else "" for n,i in enumerate(tone_sheet[target_loc_index][4:44:3])][10:]
tone_class_mark = [i.value if i.value else "" for n,i in enumerate(tone_sheet[target_loc_index][5:44:3])]
tone_class_mark[6] = "1" if tone_class_mark[6]=="" else tone_class_mark[6]
tone_class_mark[7] = "3" if tone_class_mark[7]=="" else tone_class_mark[7]
tone_class_mark[8] = "6" if tone_class_mark[8]=="" else tone_class_mark[8]
#tone_class_mark[6:9] = np.where(np.array(tone_class_mark[6:9])=="", ["1", "3", "6"], tone_class_mark[6:9])
tone_class_mark = [str(i) if i!='' else str(n+1) for n,i in enumerate(tone_class_mark)]
tone_class_markraw = [str(n+1) for n in range(len(tone_class_mark))]
tone_class_markraw_to_mark = {i[0]:i[1] for i in zip(tone_class_markraw, tone_class_mark)}
tone_class_markraw_to_contour = {i[0]:i[1] for i in zip(tone_class_markraw, tone_class_contour)}
# 對鬱林大塘：
# tone_class_mark ==> ['1', '2', '3', '4', '5', '6', '1', '2', '4', '6', '01', '02', '03']
# tone_class_markraw ==> ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13']
# tone_class_markraw_to_mark ==> {'1':'1', ..., '7':'1', ..., '11':'01', ...}
# tone_class_markraw_to_contour ==> {'1':'454', ..., '7':'5', ...}

tone_config = DictConfig(content={})
tone_config.有下陰入 = tone_class_markraw_to_contour["8"]!=""
tone_config.有下陽入 = tone_class_markraw_to_contour["10"]!=""
tone_config.全濁上變去 = config_全濁上變去
tone_config.次濁上變去 = config_次濁上變去
tone_config.應用調號 = tone_class_markraw_to_mark
tone_config.應用調號[""] = "" # 字表缺調時
tone_config.次濁聲紐 = "日來明泥娘疑以云" # 微
tone_config.全濁聲紐 = "並定從邪澄崇常船群匣"

#-----------------------------------------#

config = DictConfig(content={"initial":initial_config, "final":final_config, "tone":tone_config})

#-----------------------------------------#

print("[4] 即將輸出")
if not os.path.isfile(output_sheet_name): # 未有生成字表則生成（帶格式），已有則讀取
    output_sheets = Workbook()
    output_sheet = output_sheets.worksheets[0]
    output_sheet.title = "總字表"
    output_sheet.column_dimensions["A"].width = 4
    for n in range(2,8):
        output_sheet.column_dimensions[get_column_letter(n)].width = 3
    output_sheet.freeze_panes = 'A2'
    for m,j in enumerate(["字","母","攝","呼","等","韻","調","推導聲母","推導韻母","推導聲調","修正聲母","修正韻母","修正聲調"]):
        output_sheet.cell(1, m+1).value = j
    output_sheet["H1"].comment = Comment("聲母底色含義如下：\n綠底可能有另讀\n藍底要注意今陰陽調可能不同\n黃底表示該音很可能是不再使用的舊音", author="EcisrH")
    output_sheet["I1"].comment = Comment("韻母底色含義如下：\n綠底可能有另讀\n黃底表示該音很可能是不再使用的\n紅底表示本字很可能已經出韻", author="EcisrH")
else:
    output_sheets = load_workbook(filename=output_sheet_name, data_only=True)
    output_sheet = output_sheets["總字表"]


row_index = 2
chara_finished_count = 0
target_charas = list(chara_to_position.keys()) # jgzw 中古音字表
for target_chara in target_charas:
    if target_chara not in chara_to_position: 
        output_sheet.cell(row_index, 1).value = target_chara; row_index+=1; continue
    target_chara_position = chara_to_position[target_chara]
    target_chara_prons = retrive_pron_by_position(target_chara_position, target_chara, config) # jgzw
    
    for n,i in enumerate(target_chara_prons):
        i = [""]*6 if len(i)==0 else i
        row = [target_chara]+target_chara_position[n]+i[:3]
        marker_i, marker_f, marker_t = i[3], i[4], i[5]
        for m,j in enumerate(row):
            cell = output_sheet.cell(row_index, m+1)
            cell.value = j
            if " " in j or "?" in j: # 轉換音有空格，高亮
                cell.font = Font(color="FF51BA", bold=True)
            if m==0: # 字頭
                cell.fill = PatternFill("solid", fgColor="DAF7F8")
                cell.font = Font(bold=True, size=14)
            if m==7 and marker_i!=0: # 推導聲母有標記
                if marker_i==1: cell.fill = PatternFill("solid", fgColor="80F07C")
                elif marker_i==2: cell.fill = PatternFill("solid", fgColor="7CEDF0")
                elif marker_i==3:cell.fill = PatternFill("solid", fgColor="F0E17C")
            if m==8 and marker_f!=0: # 推導韻母有標記
                if marker_f==1: cell.fill = PatternFill("solid", fgColor="80F07C")
                elif marker_f==2: cell.fill = PatternFill("solid", fgColor="F0E17C")
                elif marker_f==3:cell.fill = PatternFill("solid", fgColor="F07C7C")
        row_index += 1
    chara_finished_count += 1
    if (chara_finished_count%500==0): print(f"已處理 {chara_finished_count} 字，共 {len(target_charas)} 字")
if row_index<=output_sheet.max_row: output_sheet.delete_rows(row_index, output_sheet.max_row)
output_sheets.save(output_sheet_name)

print("[5] 完成")